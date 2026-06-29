#!/usr/bin/env python3
"""
Conversor: Excel/planilla (formato plantilla del catalogo) -> parts.json

Uso:
    python3 tools/excel_to_parts.py [origen.xlsx] [destino.json] [hoja]

Por defecto: Nexolibre-Catalogo-Bobinas-Ago2025.xlsx -> parts.json (hoja "Catalogo")

Funciona con cualquier planilla cuya fila 1 tenga los encabezados:
ref, nombre, categoria, modalidad, marca, modelo_compatible, nro_parte,
estado, disponibilidad, ubicacion, garantia, precio, descripcion, imagen, link_externo
(tambien sirve para un export de la Lista de SharePoint a Excel/CSV).
"""
import sys, json, datetime
import openpyxl

KEYS = ["ref","nombre","categoria","modalidad","marca","modelo_compatible","nro_parte",
        "estado","disponibilidad","ubicacion","garantia","precio","descripcion","imagen","link_externo"]

def norm(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    return str(v).strip()

def main():
    src   = sys.argv[1] if len(sys.argv) > 1 else "Nexolibre-Catalogo-Bobinas-Ago2025.xlsx"
    out   = sys.argv[2] if len(sys.argv) > 2 else "parts.json"
    sheet = sys.argv[3] if len(sys.argv) > 3 else "Catalogo"

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [norm(c.value).lower() for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(len(headers))]
        if all(v in (None, "") for v in vals):
            continue
        rec = {}
        for k in KEYS:
            i = idx.get(k)
            rec[k] = norm(vals[i]) if (i is not None and i < len(vals)) else ""
        if not rec.get("nombre") and not rec.get("ref"):
            continue
        rows.append(rec)

    with open(out, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"OK: {len(rows)} piezas -> {out}")

if __name__ == "__main__":
    main()
