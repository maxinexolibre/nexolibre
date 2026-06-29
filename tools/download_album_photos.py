#!/usr/bin/env python3
"""
Descarga las fotos de los álbumes de Google Photos enlazados en
'Registro de bobinas.xlsx', las OPTIMIZA (redimensiona + comprime) y las
agrega al catálogo (parts.json) — todo en una corrida.

Mapea ID del Excel == Cód. Nexo (en la descripción de cada pieza).
- Saltea piezas que ya tengan fotos en assets/parts/.
- Baja hasta MAX_FOTOS por pieza, con pausas para no gatillar el bloqueo de Google.
- Al final corre optimize_images.py y wire_images.py.
"""
import openpyxl, json, re, os, subprocess, time

HERE      = os.path.dirname(os.path.abspath(__file__))
XLSX      = "/Users/Maxi/Downloads/Registro de bobinas.xlsx"
PARTS_DIR = "assets/parts"
MAX_FOTOS = 6
SIZE      = "=w1600"
UA        = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
COOKIES   = "/tmp/gp_cookies.txt"

def curl(extra):
    return subprocess.run(["curl","-s","--max-time","40","-A",UA,"-c",COOKIES,"-b",COOKIES]+extra,
                          capture_output=True, text=True)

def resolve(url):
    if "goo.gl" not in url:
        return url
    for line in curl(["-IL", url]).stdout.splitlines():
        if line.lower().startswith("location:"):
            return line.split(":",1)[1].strip()
    return url

def fetch(url):
    return curl(["-L", resolve(url)]).stdout

def download(url, path):
    r = curl(["-L", url, "-o", path, "-w", "%{http_code}"])
    return (r.stdout.strip().endswith("200")) and os.path.exists(path) and os.path.getsize(path) > 3000

def main():
    os.makedirs(PARTS_DIR, exist_ok=True)
    ws = openpyxl.load_workbook(XLSX)["Laboratorio"]
    id2link = {}
    for r in range(2, ws.max_row+1):
        idv = ws.cell(r,1).value; c = ws.cell(r,21)
        if idv is not None and c.hyperlink:
            try: id2link[int(idv)] = c.hyperlink.target
            except: pass

    data = json.load(open("parts.json"))
    cache = {}; total = 0
    for p in data:
        ref = (p.get("ref","") or "").strip(); rl = ref.lower()
        m = re.search(r'Cód\.\s*Nexo\s*(\d+)', p.get("descripcion",""))
        link = id2link.get(int(m.group(1))) if m else None
        if not link:
            continue
        existing = [f for f in os.listdir(PARTS_DIR)
                    if f.lower().startswith(rl+"-") or os.path.splitext(f)[0].lower()==rl]
        if existing:
            print(f"{ref}: ya tiene fotos, salteo"); continue
        if link not in cache:
            cache[link] = fetch(link); time.sleep(2.0)
        urls = []
        for u in re.findall(r'https://lh3\.googleusercontent\.com/pw/[A-Za-z0-9_-]+', cache[link]):
            if u not in urls: urls.append(u)
        if not urls:
            print(f"{ref}: sin fotos accesibles (Google bloqueó o álbum privado)"); continue
        n = 0
        for i,u in enumerate(urls[:MAX_FOTOS],1):
            if download(u+SIZE, os.path.join(PARTS_DIR, f"{rl}-{i}.jpg")):
                n += 1; total += 1
            time.sleep(0.4)
        print(f"{ref}: {n} foto(s) descargada(s)")

    print(f"\nTOTAL fotos nuevas: {total}")
    print("\n== Optimizando imágenes ==")
    subprocess.run(["python3", os.path.join(HERE, "optimize_images.py")])
    print("\n== Actualizando catálogo (parts.json) ==")
    subprocess.run(["python3", os.path.join(HERE, "wire_images.py")])

if __name__ == "__main__":
    main()
